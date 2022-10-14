import pandas as pd
import numpy as np
import os
import openpyxl
import dateutil
import datetime
from datetime import timedelta
import zipfile
from math import ceil
from pathlib import Path
from scipy import interpolate

#####################################################################################################
#                           TEXT AND DATA REPLACEMENT / INVALID DATA CLEANING                       #
#####################################################################################################
dic_tmi = pd.DataFrame(np.array([
    ['ENGINE SPEED','RPM'],
    ['ENGINE POWER','POWER'],
    ['PERCENT LOAD','LOAD'],
    ['ENGINE LOAD','LOAD'],
    ['ENGINE TORQUE','TORQUE'],
    ['BRAKE MEAN EFF PRES','BMEP'],
    ['BRAKE SPEC FUEL CONSUMPTN','BSFC'],
    ['ISO BRAKE SPEC FUEL CONSUMPTN','ISOBSFC'],
    ['VOL FUEL CONSUMPTN','VFC'],
    ['ISO VOL FUEL CONSUMPTN','ISOVFC']
    ]
),columns=['de','para'])

events_replace_text = pd.DataFrame(np.array(
    [['Severity', 'Severity'], ['Type', 'Type'], ['Code', 'Code'], ['Source', 'Source'], ['Description', 'Description'],
     ['Sample Time', 'Timestamp'], ['Run Hours', 'Run_Hours'], ['Events', 'Event'], ['Disgnostics', 'Disgnostic']]
), columns=['de', 'para'])

data_replace_text = pd.DataFrame(np.array([
    ['Sample', 'Timestamp'], ['Engine Load', 'Load'], ['Engine Speed', 'RPM'],
    ['Engine Coolant', 'Coolant_Temp'], ['Oil Pressure', 'Oil_Press'], ['Oil Temperature', 'Oil_Temp'],
    ['Battery Voltage', 'Batt'],
    ['Boost Pressure', 'Boost'], ['Fuel Consumption Rate', 'Fuel_Rate'], ['Left Exhaust Temp', 'EXH_L'],
    ['Right Exhaust Temp', 'EXH_R'],
    ['Total Fuel', 'Total_Fuel'], ['Run Hours', 'SMH'], ['Fuel Pressure', 'Fuel_Press'],
    ['Crankcase Pressure', 'Crank_Press'],
    ['Latitude', 'Latitude'], ['Longitude', 'Longitude'], ['Vessel Speed', 'Vessel_Speed'],
    ['Diagnostic Status', 'DIAG_STATS'], ['Diagnostic Code - CID', 'CID'], ['Diagnostic SubCode - FMI', 'FMI'],
    ['Event Status', 'EVENT_STATS'], ['Event Code', 'EID']
]
), columns=['de', 'para'])

maintenance_list = ['Site', 'Asset', 'SH_by_Day', 'SMH_Calc', 'SMH', 'Fuel_by_Day', 'Total_Fuel_Calc', 'Total_Fuel',
                    'Next_Prev', 'Prev_Date', 'Next_OVH', 'OVH_Date', 'Next_OVH_FC', 'OVH_FC_Date']

std_param_list = ['Timestamp', 'Load', 'RPM', 'Coolant_Temp', 'Oil_Press', 'Oil_Temp', 'Batt', 'Boost', 'Fuel_Rate',
                  'EXH_L', 'EXH_R', 'Total_Fuel', 'Fuel_Press', 'Crank_Press', 'Total_Fuel_DIFF', 'SMH_DIFF',
                  'EXH_DIFF']

std_event_list = ['Severity', 'Type', 'Code', 'Source', 'Description', 'Timestamp', 'Run_Hours']

std_eventsum_list = ['Severity', 'Type', 'Code', 'Source', 'Description']

std_rpmhist_list = ['RPM Range (%)', 'Time (%)', 'Hours (h)', 'Asset', 'Site']

std_loadhist_list = ['Power Range (%)', 'Time (%)', 'Hours (h)', 'Asset', 'Site']

remove_prefix_list = ['dg1_', 'dg2_', 'dg3_', 'dg4_', 'ple_cat_', 'mca', 'mcp', 'bb', 'be', 'cn', '_']

list_invalid_data = [-16384, 16384, -2147483648, 2147483647, 65535, -255, 255, 1016138259.02378, 1016138259, 4294967295]

#####################################################################################################
#                                           BASIC FUNCTIONS START                                   #
#####################################################################################################
def findpower(df,rpmval):
    x = df['RPM']
    y = df['POWER']
    f = interpolate.interp1d(x,y,fill_value='extrapolate')
    pwval = f(rpmval)
    return pwval

def powercalc(tmidf,engdf):
    df = replace_coltext_df(tmidf, dic_tmi)#.apply(pd.to_numeric, errors='ignore')

    for col in engdf.columns:
        if col != 'Timestamp':
          engdf[col] = pd.to_numeric(engdf[col], errors='ignore')

    for col in df.columns:
        if col != 'Timestamp':
          df[col] = pd.to_numeric(df[col], errors='ignore')

    if not 'LOAD' in tmidf.columns:
        
        if len(engdf['RPM']) > 1:
            engdf.drop(engdf[(engdf['RPM'] == 0.0) | (engdf['Load'] == 0.0) | 
                                (engdf['RPM'] < 0.0) | (engdf['Load'] < 0.0) | 
                                (engdf['RPM'] >2500) | (engdf['Load'] > 150)
                                ].index,inplace=True)
            engdf['MaxPower'] = findpower(df,engdf['RPM'])
            engdf['RealPower']=engdf['MaxPower']*engdf['Load']/100
            engdf['BSFC'] = engdf['Fuel_Rate']/engdf['RealPower']*f_density
            engdf['BSFC'] = engdf['BSFC'].apply(pd.to_numeric)
    else:
        if len(engdf['Load']) > 1:
            engdf.drop(engdf[(engdf['RPM'] == 0.0) | (engdf['Load'] == 0.0) | 
                                (engdf['RPM'] < 0.0) | (engdf['Load'] < 0.0) | 
                                (engdf['RPM'] >2500) | (engdf['Load'] > 150)
                                ].index,inplace=True)
            engdf['MaxPower'] = df['POWER'].max()
            engdf['RealPower'] = engdf['MaxPower']*engdf['Load']/100
            engdf['BSFC'] = engdf['Fuel_Rate']/engdf['RealPower']*f_density
            engdf['BSFC'] = engdf['BSFC'].apply(pd.to_numeric)
    return engdf

def openfilewb(input_file,ws):
    in_file = os.path.join(input_file)
    outwb = openpyxl.load_workbook(in_file)
    data = outwb[ws].values
    columns = next(data)[0:]
    df = pd.DataFrame(data,columns=columns)
    df.dropna(how='all',inplace=True)
    return df 

def cons_perfn(asset):
    perfseries = assetlistdf.loc[assetlistdf['Serial'] == asset]['Perf_Number'].tolist()
    return perfseries[0]

def limpadao(dir):
    for stuff in os.listdir(dir):
        coisa = os.path.join(dir, stuff)
        if os.path.isdir(coisa) == True:
            for stuff2 in os.listdir(coisa):
                if os.path.isdir(stuff2) == False:
                    if stuff2.endswith('output.csv'):
                        os.replace(os.path.join(coisa, stuff2), os.path.join(dir, stuff2))
    if keepfiles == 0:
        for file_name in os.listdir(dir):
            folder = os.path.join(dir, file_name)
            if os.path.isdir(folder) == True:
                os.rmdir(folder)

# Tenta criar o diretório. Caso já exista, passa adiante.
def checkdestiny(outputdirectory):
    try:
        os.mkdir(outputdirectory)
    except FileExistsError:
        pass

def removeprefix(text, plist):
    for ch in plist:
        if ch in text:
            text = text.replace(ch, "")
    return text

def replace_coltext_df(df, replacement_df):
    o = len(replacement_df.axes[0])
    r = len(replacement_df.axes[1])
    i = 0
    while i < o:
        df.columns = df.columns.str.replace(r'('+replacement_df.iloc[i, 0] + '.*$)', replacement_df.iloc[i, 1],regex=True)
        i += 1
    return df

def replace_text(df, replacement_df):
    o = len(replacement_df.axes[0])
    r = len(replacement_df.axes[1])
    i = 0
    while i < o:
        df.columns = df.columns.str.replace(r'(^.*' + replacement_df.iloc[i, 0] + '.*$)', replacement_df.iloc[i, 1],
                                            regex=True)
        i += 1
    return df

def csvfix(df, lista):
    dffix = df.copy(deep=True)
    for h in lista:
        if h not in dffix.columns:
            dffix[h] = ''
    return dffix

def csvfix2(colunas, linhas):
    dffix = pd.DataFrame(linhas, columns=colunas)
    return dffix

def dateparser(filein):
    date_parser = lambda x: dateutil.parser.parse(x, ignoretz=True)
    tabin = pd.read_csv(filein, dayfirst=False, parse_dates=['Sample Time'], date_parser=date_parser)
    tabin.to_csv(filein, encoding='utf-8-sig', index=True)

def getlistativos(file):
    # pega a aba 'Engine Event Summary' e cria um array com todos os números de série da coluna 'Unit Name', tirando a linha de totais.
    if file.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file)
        global ws_Eng_Summ
        ws_Eng_Summ = wb['Engine Event Summary']
        data = ws_Eng_Summ.values
        df = pd.DataFrame(data, columns=next(data)[0:])
        df = df.loc[:, df.columns == 'Unit Name']
    return df['Unit Name']

def getlistasites(file):
    # pega a aba 'ASSET_LIST' e cria um array com todos os nomes de sítios.
    if file.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file)
        global ws_Site_Summ
        ws_Site_Summ = wb['ASSET_LIST']
        data = ws_Site_Summ.values
        df = pd.DataFrame(data, columns=next(data)[0:])
        df = df.loc[:, df.columns == 'Vessel']
        df.drop_duplicates(subset=['Vessel'],inplace=True)
    return df['Vessel']

def concatenar(dir, name):
    outname = name + 'output.csv'
    for file_name in os.listdir(dir):
        if file_name != outname:
            if file_name != scriptname:
                a = pd.read_csv(dir + '/' + outname, low_memory=False)
                b = pd.read_csv(dir + file_name, low_memory=False)
                a.columns = a.columns.str.replace(" ","_",regex=True)
                b.columns = b.columns.str.replace(" ","_",regex=True)
                df = pd.concat([a, b])
                df.drop_duplicates(subset=['Timestamp', 'Asset'], inplace=True, keep='last')
                # df.dropna(how='all', axis=1, inplace=True)
                df.to_csv(dir + '/' + outname, encoding='utf-8-sig', index=False)
    if keepfiles == 0:
        for file_name in os.listdir(dir):
            if file_name != outname:
                os.remove(dir + file_name)
    return df

def concatenarev(dir, name):
    outname = name + 'output.csv'
    for file_name in os.listdir(dir):
        if file_name != outname:
            if file_name != scriptname:
                a = pd.read_csv(dir + '/' + outname, low_memory=False)
                b = pd.read_csv(dir + file_name, low_memory=False)
                a.columns = a.columns.str.replace(" ","_",regex=True)
                b.columns = b.columns.str.replace(" ","_",regex=True)
                df = pd.concat([a, b])
                df.drop_duplicates(inplace=True, keep='last')
                # df.dropna(how='all', axis=1, inplace=True)
                df.to_csv(dir + '/' + outname, encoding='utf-8-sig', index=False)
                
    if keepfiles == 0:
        for file_name in os.listdir(dir):
            if file_name != outname:
                os.remove(dir + file_name)

    return df

def concatenar_profile(hdir, columns, name):
    outname = name + 'output.csv'
    try:
        os.remove(hdir + '/' + outname)
        cdf1 = pd.DataFrame(columns=columns)
        cdf1.to_csv(hdir + '/' + outname, encoding='utf-8-sig', index=False)
    except FileNotFoundError:
        cdf1 = pd.DataFrame(columns=columns)
        cdf1.to_csv(hdir + '/' + outname, encoding='utf-8-sig', index=False)

    for file_name in os.listdir(hdir):
        if file_name != outname:
            if file_name != scriptname:
                a = pd.read_csv(hdir + '/' + outname, low_memory=False)
                b = pd.read_csv(hdir + file_name, low_memory=False)
                dfs = pd.concat([a, b])
                dfs.drop_duplicates(subset=None, inplace=True, keep='last')
                # dfs.dropna(how='all', axis=1, inplace=True)
                dfs.to_csv(hdir + '/' + outname, encoding='utf-8-sig', index=False)
    if keepfiles == 0:
        for file_name in os.listdir(hdir):
            if file_name != outname:
                os.remove(hdir + file_name)
    return dfs

def concatenar_study(hdir, name):
    outname = name + 'output.csv'
    try:
        os.remove(hdir + '/' + outname)
        cdf = pd.DataFrame(columns=['DataM'])
        cdf.to_csv(hdir + '/' + outname, encoding='utf-8-sig', index=False)
    except FileNotFoundError:
        cdf = pd.DataFrame(columns=['DataM'])
        cdf.to_csv(hdir + '/' + outname, encoding='utf-8-sig', index=False)

    for file_name in os.listdir(hdir):
        if file_name != outname:
            if file_name != scriptname:
                a = pd.read_csv(hdir + '/' + outname, low_memory=False)
                b = pd.read_csv(hdir + file_name, low_memory=False)
                df = pd.concat([a, b])
                df.drop_duplicates(subset=None, inplace=True, keep='last')
                # df.dropna(how='all', axis=1, inplace=True)
                try:
                    df.drop(columns=['DataM'],inplace=True)
                except KeyError:
                    pass
                df.to_csv(hdir + '/' + outname, encoding='utf-8-sig', index=False)
                
    if keepfiles == 0:
        for file_name in os.listdir(hdir):
            if file_name != outname:
                os.remove(hdir + file_name)

    return df

def findsitename(asset):
    wb_info = openpyxl.load_workbook(os.path.join(infodir, 'ASSET_INFO.xlsx'))
    ws_a_list = wb_info['ASSET_LIST']
    data = ws_a_list.values
    cuslist = pd.DataFrame(data, columns=next(data)[0:])
    site_name = cuslist.loc[cuslist['Serial'] == asset]['Vessel'].tolist()
    return site_name[0]

def findsiteassets(site):
    wb_info = openpyxl.load_workbook(os.path.join(infodir, 'ASSET_INFO.xlsx'))
    ws_a_list = wb_info['ASSET_LIST']
    data = ws_a_list.values
    df = pd.DataFrame(data, columns=next(data)[0:])
    slist = df.loc[df['Vessel'] == site]['Serial'].tolist()
    return slist


#####################################################################################################
#                                           MANUTENÇÃO                                              #
#####################################################################################################
# ele vai ter que receber o nome do ativo para filtrar no dataframe esse ativo 
# e aí fazer os cálculos de manutenção!!!!

def maintenanceoutput(dfoutput, lastused, asset_sn, datasetvazio):
    site_name = findsitename(asset_sn)

    fcd, fct, totfc = fuelcalc(dfoutput, asset_sn)
    shd, sht, totsh = smhcalc(dfoutput, asset_sn)

    print('Ativo:', str(asset_sn))
    print('Consumo por dia:', str(fcd), 'Consumo Total:', str(fct))
    print('Horas de serviço por dia:', str(shd), 'Horimetro:', str(sht))

    next_preventiva, next_preventiva_day = manutcalc(sht, shd, lastused, asset_sn, datasetvazio)
    next_overhaul, next_overhaul_day, next_overhaul_fc, next_overhaul_fc_day = ovhaulcalc(sht, shd, fct, fcd, lastused,
                                                                                          asset_sn, datasetvazio)

    print('Proxima manutencao preventiva:', str(next_preventiva), 'horas.', 'Data:', str(next_preventiva_day))
    print('Proximo overhaul:', str(next_overhaul), 'horas.', 'Data:', str(next_overhaul_day))
    print('Proximo overhaul (consumo):', str(next_overhaul_fc), 'litros.', 'Data:', str(next_overhaul_fc_day))
    print(' ')

    df = pd.DataFrame([[site_name, asset_sn, shd, sht, totsh, fcd, fct, totfc, next_preventiva, next_preventiva_day,
                        next_overhaul, next_overhaul_day, next_overhaul_fc, next_overhaul_fc_day]],
                      columns=maintenance_list)
    return df

def fuelcalc(df, assetname):
    global fuelbyday
    global fuelcons
    global totalfuel
    global ndays

    df = df.query("Asset == @assetname")
    df = df.query("RPM > 0")
    if not df.shape[0] < 1:
        if 'Total_Fuel' in df.columns:
            try:
                ndays = np.timedelta64(pd.to_datetime(df['Timestamp']).max()-pd.to_datetime(df['Timestamp']).min(),'h').astype(int)/24
            except:
                ndays = 1

            totalfuel = round(pd.to_numeric(df['Total_Fuel'].max()), 0)
            try:
                fuelcons = pd.to_numeric(df['Total_Fuel'].max() - df['Total_Fuel'].min())
                fuelcons = round(fuelcons, 0)
            except:
                fuelcons = np.nan

            if np.isnan(fuelcons) or ndays == 1 or np.isnan(ndays) or ndays == 0 or fuelcons == 0:
                fuelcons = np.nan
                fuelbyday = np.nan
            else:
                try:
                    fuelbyday = int(int(fuelcons) / ndays)
                    fuelbyday = round(fuelbyday, 0)
                except ValueError:
                    fuelbyday = np.nan
    else:
        fuelcons = np.nan
        fuelbyday = np.nan
        totalfuel = np.nan
    return [fuelbyday, fuelcons, totalfuel]

def smhcalc(df, assetname):
    global smhbyday
    global servicehr
    global totalsh
    global ndays

    df = df.query("Asset == @assetname")
    df = df.query("RPM > 0")
    if not df.shape[0] < 1:
        if 'SMH' in df.columns:
            try:
                ndays = np.timedelta64(pd.to_datetime(df['Timestamp']).max() - pd.to_datetime(df['Timestamp']).min(),
                                   'h').astype(int) / 24
            except:
                ndays = 1

            totalsh = round(pd.to_numeric(df['SMH'].max()), 0)
            try:
                servicehr = pd.to_numeric(df['SMH'].max() - df['SMH'].min())
                servicehr = round(servicehr, 0)
            except:
                servicehr = np.nan

            if np.isnan(servicehr) or ndays == 1 or np.isnan(ndays) or ndays == 0 or servicehr == 0:
                servicehr = np.nan
                smhbyday = np.nan
            else:
                try:
                    smhbyday = ceil(servicehr / ndays)
                    smhbyday = round(smhbyday, 0)
                except ValueError:
                    smhbyday = np.nan
    else:
        servicehr = np.nan
        smhbyday = np.nan
        totalsh = np.nan
    return [smhbyday, servicehr, totalsh]

# next_preventiva, next_preventiva_day = manutcalc(sht, shd, lastused, asset_sn, datasetvazio)
def manutcalc(fhor, hday, lastused, sn, datasetvazio):
    if datasetvazio == 1:
        nextprev = 'Ativo sem dados'
        mandday = np.nan
        return [nextprev, mandday]
    else:
        # PEGA O PREFIXO DE SÉRIE PARA BUSCAR NA PLANILHA ONDE ESTÃO OS INTERVALOS DE MANUTENÇÃO
        sn_pref = sn[-8:3]
        try:
            manplan = pd.read_csv(os.path.join(infodir + '/MAINTENANCE_PLAN.csv'))
            # manplan = manplan.apply(pd.to_numeric, errors='coerce')
            manplan = manplan.dropna(how='all')
        except FileNotFoundError:
            nextprev = 'No MAINTENANCE_PLAN.csv file'
            mandday = np.nan
            return [nextprev, mandday]

        # print('***interno função manutenção***')    
        manplan = manplan.apply(pd.to_numeric, errors='ignore')
        fhor = float(fhor)
        # print('Horímetro do motor: ' + str(fhor))
        hday = float(hday)
        # print('Horas por dia: ' + str(hday))
        lastused = pd.to_datetime(lastused)
        # print('Usado pela última vez: ' + str(lastused))

        try:
            manshift = pd.read_csv(os.path.join(infodir + '/MAINTENANCE_SHIFT.csv'))
            manshift.columns = manshift.columns.str.lower()
            manshift.columns = manshift.columns.str.replace(' - ', '_')
            # manutenção realizada
            forcedman = float(manshift.loc[0, sn])
            # print(forcedman)
            if np.isnan(forcedman):
                pass
            else:
                # dia da realização
                fmandate = pd.to_datetime(manshift.loc[1, sn])
                # print(fmandate)
                # diferença do horímetro
                #hordiff = abs(forcedman - fhor)
                # print(hordiff)
                # dias decorridos desde então
                shiftdays = lastused - fmandate
                shiftdays = int(shiftdays.days)
                # print(shiftdays)
                # horas rodadas desde a ultima revisão
                horsinceman = shiftdays * hday
                # print(horsinceman)
                # diff horímetro simulado
                horshift = float(horsinceman) + float(forcedman)
                # print(horshift)
                #fhor = fhor + horshift <------------------ ok
                fhor = horshift
                # print(fhor)
        except KeyError:
            pass
        except FileNotFoundError:
            pass

        # print(sn_pref)

        colprev_h = sn_pref + '_PSH'
        colprev_l = sn_pref + '_PFC'

        # by service hours - periódica
        if np.isnan(fhor) or np.isnan(hday):
            nextprev = np.nan
            mandday = np.nan
        else:
            lastman = manplan[colprev_h].max()
            # print('Manutenção mais alta do plano: ' + str(lastman))
            # número de loops
            nloops = fhor / lastman
            # print('Número de Loops: '+ str(nloops))
            # difereça de descarte
            ndesc = int(nloops) * lastman
            # print('Descarte: ' +str(ndesc))
            # horimetro restante
            horest = fhor - ndesc
            # print('Horímetro restante: ' +str(horest))
            # proxima manutenção
            # print(colprev_h)
            # print(manplan[colprev_h])
            cplan = [x for x in manplan[colprev_h] if str(x) != 'nan']
            # print(cplan)
            npindex = pd.Index(cplan).dropna().get_indexer([horest], method='bfill')[0]
            nextprev = manplan[colprev_h].iloc[npindex]
            # print('Próxima revisão: ' +str(nextprev))
            # quantos dias para proxima revisão
            ndiasrev = (nextprev - horest) / hday
            # print('Dias para revisão: '+ str(ndiasrev))
            # data da proxima manutenção
            mandday = (lastused + timedelta(days=ndiasrev)).round(freq='D').date()
            # print('Data da próxima revisão: ' +str(mandday))
        # print('***interno função manutenção***')
        return [nextprev, mandday]

def ovhaulcalc(fhor, hday, fcons, fday, lastused, sn, datasetvazio):
    if datasetvazio == 1:
        nextovhaul = 'Ativo sem dados'
        fc_nextovhaul = 'Ativo sem dados'
        ovhauldday = np.nan
        fc_ovhauldday = np.nan
        return [nextovhaul, ovhauldday, fc_nextovhaul, fc_ovhauldday]
    else:

        sn_pref = sn[-8:3]
        try:
            manplan = pd.read_csv(os.path.join(infodir + '/MAINTENANCE_PLAN.csv'))
            # manplan = manplan.apply(pd.to_numeric, errors='coerce')
            manplan = manplan.dropna(how='all')
        except FileNotFoundError:
            nextovhaul = 'No MAINTENANCE_PLAN.csv file'
            fc_nextovhaul = 'No MAINTENANCE_PLAN.csv file'
            ovhauldday = np.nan
            fc_ovhauldday = np.nan
            return [nextovhaul, ovhauldday, fc_nextovhaul, fc_ovhauldday]

        colovhaul_h = sn_pref + '_OSH'
        colovhaul_l = sn_pref + '_OFC'

        # ---------------------------HORAS DE SERVIÇO-------------------------

        # print('***interno função manutenção***')    
        manplan = manplan.apply(pd.to_numeric, errors='ignore')
        fhor = float(fhor)
        # print('Horímetro do motor: ' + str(fhor))
        hday = float(hday)
        # print('Horas por dia: ' + str(hday))
        lastused = pd.to_datetime(lastused)
        # print('Usado pela última vez: ' + str(lastused))

        try:
            manshift = pd.read_csv(os.path.join(infodir + '/MAINTENANCE_SHIFT.csv'))
            manshift.columns = manshift.columns.str.lower()
            manshift.columns = manshift.columns.str.replace(' - ', '_')
            # manutenção realizada
            forcedman = float(manshift.loc[0, sn])
            # print(forcedman)
            if np.isnan(forcedman):
                pass
            else:
                # dia da realização
                fmandate = pd.to_datetime(manshift.loc[1, sn])
                # print(fmandate)
                # diferença do horímetro
                #hordiff = abs(forcedman - fhor)
                # print(hordiff)
                # dias decorridos desde então
                shiftdays = lastused - fmandate
                shiftdays = int(shiftdays.days)
                # print(shiftdays)
                # horas rodadas desde a ultima revisão
                horsinceman = shiftdays * hday
                # print(horsinceman)
                # diff horímetro simulado
                horshift = float(horsinceman) + float(forcedman)
                # print(horshift)
                #fhor = fhor + horshift <---------------------- ok
                fhor = horshift
                # print(fhor)
        except KeyError:
            pass
        except FileNotFoundError:
            pass

        if np.isnan(fhor) or np.isnan(hday):
            nextovhaul = np.nan
            ovhauldday = np.nan
        else:
            lastman = manplan[colovhaul_h].max()
            # print('Manutenção mais alta do plano: ' + str(lastman))
            # número de loops
            nloops = fhor / lastman
            # print('Número de Loops: '+ str(nloops))
            # difereça de descarte
            ndesc = int(nloops) * lastman
            # print('Descarte: ' +str(ndesc))
            # horimetro restante
            horest = fhor - ndesc
            # print('Horímetro restante: ' +str(horest))
            # proxima manutenção
            npindex = pd.Index(manplan[colovhaul_h]).dropna().get_indexer([horest], method='bfill')[0]
            nextovhaul = manplan[colovhaul_h].iloc[npindex]
            # print('Próxima revisão: ' +str(nextovhaul))
            # quantos dias para proxima revisão
            ndiasrev = (nextovhaul - horest) / hday
            # print('Dias para revisão: '+ str(ndiasrev))
            # data da proxima manutenção
            ovhauldday = (lastused + timedelta(days=ndiasrev)).round(freq='D').date()
            # print('Data da próxima revisão: ' +str(ovhauldday))
        # print('***interno função manutenção***')

        # ---------------------------CONSUMO DE COMBUSTÍVEL-------------------------

        # print('***interno função manutenção***')    
        manplan = manplan.apply(pd.to_numeric, errors='ignore')
        fcons = float(fcons)
        # print('Horímetro do motor: ' + str(fcons))
        fday = float(fday)
        # print('Horas por dia: ' + str(fday))

        try:
            manshift = pd.read_csv(os.path.join(infodir + '/MAINTENANCE_SHIFT.csv'))
            manshift.columns = manshift.columns.str.lower()
            manshift.columns = manshift.columns.str.replace(' - ', '_')
            forcedman = float(manshift.loc[2, sn])

            if np.isnan(forcedman):
                pass
            else:
                fmandate = pd.to_datetime(manshift.loc[3, sn])
                #hordiff = abs(forcedman - fcons)
                shiftdays = lastused - fmandate
                shiftdays = int(shiftdays.days)
                horsinceman = shiftdays * fday
                horshift = float(horsinceman) + float(forcedman)
                #fhor = fhor + horshift <---------------------- ok
                fcons = horshift
                
        except KeyError:
            pass
        except FileNotFoundError:
            pass

        if np.isnan(fcons) or np.isnan(fday):
            fc_nextovhaul = np.nan
            fc_ovhauldday = np.nan
        else:
            lastman = manplan[colovhaul_l].max()
            nloops = fcons / lastman
            ndesc = int(nloops) * lastman
            horest = fcons - ndesc
            npindex = pd.Index(manplan[colovhaul_l]).dropna().get_indexer([horest], method='bfill')[0]
            fc_nextovhaul = manplan[colovhaul_l].iloc[npindex]
            ndiasrev = (fc_nextovhaul - horest) / fday
            try:
                fc_ovhauldday = (lastused + timedelta(days=ndiasrev)).round(freq='D').date()
            except:
                fc_ovhauldday = np.nan
        return [nextovhaul, ovhauldday, fc_nextovhaul, fc_ovhauldday]

#####################################################################################################
#                                           HISTOGRAMAS                                             #
#####################################################################################################

def load_histogram(dataframe, rawdf, a_sn):
    perc = []
    PMax = 110
    Pdiv = 10
    h = (PMax // Pdiv) + 1
    n = 0
    for b in range(h):
        perc.append(n)
        n += 10
    dataframe["Load"] = dataframe["Load"].apply(pd.to_numeric)
    dataframe["RPM"] = dataframe["RPM"].apply(pd.to_numeric)
    engine_pw = dataframe.query("RPM > 0")["Load"]
    serieshist1 = engine_pw.value_counts(sort=False, bins=perc, normalize=True).round(4) * 100
    dfhist1 = serieshist1.rename_axis('unique_values').reset_index(name='counts')
    dfhist1 = engine_pw.value_counts(sort=False, bins=perc, normalize=True).reset_index().rename(
        columns={'index': 'bin'})
    dfhist1['Power Range (%)'] = ['({}, {})'.format(x.left, x.right) for x in dfhist1['bin']]
    dfhist1.columns = dfhist1.columns.str.replace('Load', 'Time (%)')
    dfhist1['Time (%)'] = dfhist1['Time (%)'].round(4) * 100
    dfhist1.drop(['bin'], axis=1, inplace=True)
    dfhist1['Power Range (%)'] = dfhist1['Power Range (%)'].map(
        {'(-0.001, 10.0)': '0-10%', '(10.0, 20.0)': '10-20%', '(20.0, 30.0)': '20-30%', '(30.0, 40.0)': '30-40%',
         '(40.0, 50.0)': '40-50%', '(50.0, 60.0)': '50-60%', '(60.0, 70.0)': '60-70%', '(70.0, 80.0)': '70-80%',
         '(80.0, 90.0)': '80-90%', '(90.0, 100.0)': '90-100%', '(100.0, 110.0)': '100-110%'}, na_action=None)

    if 'SMH' in rawdf.columns:
        rawdf["SMH"] = rawdf["SMH"].apply(pd.to_numeric)
        lasth = rawdf['SMH'].max()
        firsth = rawdf['SMH'].min()
        periodh = lasth - firsth
        dfhist1['Hours (h)'] = dfhist1['Time (%)'] * periodh / 100
        dfhist1 = dfhist1[['Power Range (%)', 'Time (%)', 'Hours (h)']]
    else:
        dfhist1 = dfhist1[['Power Range (%)', 'Time (%)']]
    outdir = os.path.join(destinationfolder, loadhistdir)
    checkdestiny(outdir)
    histout = outdir + a_sn + '_LOADPROFILE' + '.csv'

    listaaa = ['Power Range (%)', 'Time (%)', 'Hours (h)']
    rowws = [['0-10%', 0, 0], ['10-20%', 0, 0], ['20-30%', 0, 0], ['30-40%', 0, 0], ['50-60%', 0, 0], ['70-80%', 0, 0],
             ['90-100%', 0, 0], ['100-110%', 0, 0]]
    ndf = csvfix(dfhist1, listaaa)

    if ndf.shape[0] < 1:
        df = csvfix2(listaaa, rowws)
        df['Asset'] = asset_sn
        df['Site'] = findsitename(asset_sn)
        df.reset_index(drop=True, inplace=True)
        # df = df.transpose()
        df.to_csv(histout, encoding='utf-8-sig', index=False)
    else:
        ndf['Asset'] = asset_sn
        ndf['Site'] = findsitename(asset_sn)
        ndf.reset_index(drop=True, inplace=True)
        # ndf = ndf.transpose()
        ndf.to_csv(histout, encoding='utf-8-sig', index=False)

def rpm_histogram(dataframe, rawdf, a_sn):
    rawdf['RPM'] = rawdf['RPM'].apply(pd.to_numeric)
    # HISTOGRAMAS DE ROTAÇÃO
    perc = []
    PMax = 1900
    Pdiv = 100
    h = 20
    n = 0
    for b in range(h):
        perc.append(n)
        n += 100
    global dfhist2
    dataframe['RPM'] = dataframe['RPM'].apply(pd.to_numeric)
    engine_rpm = dataframe.query("RPM > 0")['RPM']
    serieshist2 = engine_rpm.value_counts(sort=False, bins=perc, normalize=True).round(4) * 100
    dfhist2 = serieshist2.rename_axis('unique_values').reset_index(name='counts')
    dfhist2 = engine_rpm.value_counts(sort=False, bins=perc, normalize=True).reset_index().rename(
        columns={'index': 'bin'})
    dfhist2['Range'] = ['({}, {})'.format(x.left, x.right) for x in dfhist2['bin']]
    dfhist2.columns = dfhist2.columns.str.replace('RPM', 'Time (%)')
    dfhist2.columns = dfhist2.columns.str.replace('Range', 'RPM Range (%)')
    dfhist2['Time (%)'] = dfhist2['Time (%)'].round(4) * 100
    dfhist2.drop(['bin'], axis=1, inplace=True)
    dfhist2['RPM Range (%)'] = dfhist2['RPM Range (%)'].map(
        {'(-0.001, 100.0)': '0-100', '(100.0, 200.0)': '100-200', '(200.0, 300.0)': '200-300',
         '(300.0, 400.0)': '300-400', '(400.0, 500.0)': '400-500', '(500.0, 600.0)': '500-600',
         '(600.0, 700.0)': '600-700', '(700.0, 800.0)': '700-800', '(800.0, 900.0)': '800-900',
         '(900.0, 1000.0)': '900-1000', '(1000.0, 1100.0)': '1000-1100', '(1100.0, 1200.0)': '1100-1200',
         '(1200.0, 1300.0)': '1200-1300', '(1300.0, 1400.0)': '1300-1400', '(1400.0, 1500.0)': '1400-1500',
         '(1500.0, 1600.0)': '1500-1600', '(1600.0, 1700.0)': '1600-1700', '(1700.0, 1800.0)': '1700-1800',
         '(1800.0, 1900.0)': '1800-1900'}, na_action=None)
    if 'SMH' in rawdf.columns:
        lasth = rawdf['SMH'].max()
        firsth = rawdf['SMH'].min()
        periodh = lasth - firsth
        dfhist2['Hours (h)'] = dfhist2['Time (%)'] * periodh / 100
        dfhist2 = dfhist2[['RPM Range (%)', 'Time (%)', 'Hours (h)']]
    else:
        dfhist2 = dfhist2[['RPM Range (%)', 'Time (%)']]

    outdir = os.path.join(destinationfolder, rpmhistdir)
    checkdestiny(outdir)
    histout = outdir + a_sn + '_RPMPROFILE' + '.csv'

    listbbb = ['RPM Range (%)', 'Time (%)', 'Hours (h)']

    df = csvfix(dfhist2, listbbb)
    df['Asset'] = asset_sn
    df['Site'] = findsitename(asset_sn)
    df.reset_index(drop=True, inplace=True)
    # df = df.transpose()
    df.to_csv(histout, encoding='utf-8-sig', index=False)

#####################################################################################################
#                            ESTUDOS DE UTILIZAÇÃO E TAXAS DE CONSUMO                               #
#####################################################################################################

def genloadxhour(df, base_path, loadtable, fratetable):
    vnames = df.columns
    for vname in vnames:
        dfvname = [x for x in df[vname] if str(x) != 'None']
        for a in range(len(dfvname)):
            if a == 0:
                ltv = pd.concat([loadtable.loc[:, loadtable.columns.isin(['Timestamp'])],
                                 loadtable.loc[:, loadtable.columns.str.contains(dfvname[a], case=False)]], axis=1)
                a += 1
            else:
                ltv = pd.concat([ltv, loadtable.loc[:, loadtable.columns.str.contains(dfvname[a], case=False)]], axis=1)

        if len(ltv.columns) != 0:
            ltv = ltv.replace(0, np.nan)
            ltv['QNT_SIM'] = ltv.apply(lambda x: x.notnull().sum(), axis='columns')

            fin = pd.concat([pd.DataFrame([i], columns=['QNT']) for i in range(len(dfvname) + 1)], ignore_index=True)
            l = []
            for i in range(len(dfvname) + 1):
                l.append(ltv[ltv['QNT_SIM'] == (i)].count()['QNT_SIM'] / len(ltv['QNT_SIM']) * 100)
            fin['%'] = l
            finfile = os.path.join(base_path) + '/' + vname + '_LOAD_RESUME' + '.csv'
            fin.to_csv(finfile, encoding='utf-8-sig', index=False)

            ltvd = ltv.resample('1D').mean(numeric_only=True).round(0)
            for x in range(len(ltv.columns) - 1):
                ltvd['UT' + str(x + 1) + 'E'] = (
                        (ltv['QNT_SIM'] == (x + 1)).resample('1D').sum().astype(int) / 24 * 100).round(1)

            ltvd.drop(['QNT_SIM'], axis=1, inplace=True)
            ltvdfile = os.path.join(base_path) + '/' + vname + '_LOAD_STUDY' + '.csv'
            ltvd.to_csv(ltvdfile, encoding='utf-8-sig', index=True)

        # transforma v1 em tabela de taxa de consumo / hora
        for a in range(len(dfvname)):
            if a == 0:
                rtv = pd.concat([fratetable.loc[:, fratetable.columns.isin(['Timestamp'])],
                                 fratetable.loc[:, fratetable.columns.str.contains(dfvname[a], case=False)]], axis=1)
                a += 1
            else:
                rtv = pd.concat([rtv, fratetable.loc[:, fratetable.columns.str.contains(dfvname[a], case=False)]],
                                axis=1)

        rtvb = rtv.copy()
        rtvb = rtvb.replace(0, np.nan)
        rtvb['AVG'] = rtvb.sum(numeric_only=True, axis=1)

        if len(rtv.columns) != 0:
            rtvfile = os.path.join(base_path) + '/' + vname + '_FR_STUDY' + '.csv'
            rtv = rtv.replace(0, np.nan)
            rtvD = rtv.resample('1D').mean(numeric_only=True).round(0)
            rtvD.to_csv(rtvfile, encoding='utf-8-sig', index=True)

            rtv['QNT_SIM'] = rtv.apply(lambda x: x.notnull().sum(), axis='columns')

            rin = pd.concat([pd.DataFrame([i], columns=['QNT']) for i in range(len(dfvname) + 1)], ignore_index=True)
            l = []
            for i in range(len(dfvname) + 1):
                l.append(rtvb[rtv['QNT_SIM'] == (i)].mean(numeric_only=True)['AVG'])
            rin['L/hr'] = l
            rinfile = os.path.join(base_path) + '/' + vname + '_FR_RESUME' + '.csv'
            rin.to_csv(rinfile, encoding='utf-8-sig', index=False)

#####################################################################################################
#                                           ROTINAS                                                 #
#####################################################################################################

def rotinas(dataframe):
    global dfclean
    dataframe['Asset'] = asset_sn
    dataframe['Site'] = findsitename(asset_sn)
    perfn = cons_perfn(asset_sn)
    print('Abrindo ativo:', str(asset_sn), '\nPerformance Number: ' + perfn)
    for col in dataframe.columns:
        if col != 'Timestamp':
            dataframe[col] = pd.to_numeric(dataframe[col], errors='ignore')

    dfclean = dataframe.replace(list_invalid_data, '')
    if not dfclean.shape[0] < 1:
        if 'SMH' in dfclean.columns:
            dfclean["SMH"] = dfclean["SMH"].apply(pd.to_numeric)
            dfclean['SMH'].interpolate(inplace=True)
            dfclean['SMH_DIFF'] = dfclean['SMH'].diff().bfill()
            dfclean['SMH_DIFF'] = abs(dfclean['SMH_DIFF'].apply(pd.to_numeric))
        else:
            dfclean['SMH_DIFF'] = np.nan

        if 'RPM' in dfclean.columns:
            MeanUR = dfclean['Timestamp'].diff().mean(numeric_only=True)
            Mintime = datetime.timedelta(minutes=10)
            if MeanUR < Mintime:
                dfclean['RPM'].interpolate(limit=1, inplace=True)
            dfclean = dfclean.dropna(axis=0, subset=['RPM'])
            dfclean = dfclean.drop(dfclean[(dfclean['RPM'] == 0)].index)
            rpm_histogram(dfclean, dataframe, asset_sn)

        if 'Load' in dfclean.columns:
            dfclean = dfclean.dropna(axis=0, subset=['Load'])
            dfclean = dfclean.drop(dfclean[(dfclean['RPM'] == 0) & (dfclean['Load'] == 0)].index)
            try:    
                tmidf = openfilewb(os.path.join(infodir, 'TMI_INFO.xlsx'),perfn)
                dfclean = powercalc(tmidf,dfclean)
            except KeyError:
                print('Sem dados de performance. Alguns calculos nao serao realizados, incluindo BSFC e Potencia Real.')
            load_histogram(dfclean, dataframe, asset_sn)

        if 'Total_Fuel' in dfclean.columns:
            dfclean = dfclean.drop(dfclean[(dfclean['Total_Fuel'] == 0)].index)
            dfclean["Total_Fuel"] = dfclean["Total_Fuel"].apply(pd.to_numeric)
            dfclean['Total_Fuel'].interpolate(inplace=True)
            dfclean['Total_Fuel_DIFF'] = dfclean['Total_Fuel'].diff().bfill()
            dfclean['Total_Fuel_DIFF'] = dfclean['Total_Fuel_DIFF'].apply(pd.to_numeric).abs()
        else:
            dfclean['Total_Fuel_DIFF'] = np.nan

        if all(pd.Series(['EXH_L', 'EXH_R']).isin(dfclean.columns)):
            dfclean["EXH_L"] = dfclean["EXH_L"].apply(pd.to_numeric)
            dfclean["EXH_R"] = dfclean["EXH_R"].apply(pd.to_numeric)
            dfclean['EXH_DIFF'] = abs(dfclean.EXH_L - dfclean.EXH_R)
        else:
            dfclean['EXH_DIFF'] = np.nan

        dfclean.set_index(pd.DatetimeIndex(dfclean['Timestamp']), inplace=True)
        print('Analise Ok.')
    else:
        print('Datalog vazio.')
    return dfclean

def rotinaseventos(outws, ws, asset_sn):
    site_name = findsitename(asset_sn)
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)
    df = replace_text(df, events_replace_text)
    df['Code'] = df['Code'].astype(str)
    df['Asset'] = asset_sn
    df['Site'] = site_name
    df.to_csv(outws, encoding='utf-8-sig', index=False)

def rotinaseventossum(outws2, ws2, asset_sn):
    site_name = findsitename(asset_sn)
    data2 = ws2.values
    columns = next(data2)[0:]
    dfs = pd.DataFrame(data2, columns=columns)
    dfs = replace_text(dfs, events_replace_text)

    if all(pd.Series(std_eventsum_list).isin(dfs.columns)):
        count_series = dfs.groupby(std_eventsum_list).size()
        eventsalerts = count_series.to_frame(name='Count').reset_index()

    else:
        eventsalerts = pd.DataFrame(columns=std_eventsum_list)

    eventsalerts['Code'] = eventsalerts['Code'].astype(str)
    eventsalerts['Asset'] = asset_sn
    eventsalerts['Site'] = site_name
    eventsalerts.to_csv(outws2, encoding='utf-8-sig', index=False)

#####################################################################################################
#                                          CONVERSÃO DE EVENTOS                                     #
#####################################################################################################

def eventsconvert(eventfile, ts_file):
    global troublefile
    global asset_sn
    global outws
    global outws2

    eventfile = os.path.join(eventfile)

    troublefile = os.path.join(ts_file)

    workfolder = os.path.join(destinationfolder, eventsdirectory)
    workfolder2 = os.path.join(destinationfolder, eventssumdirectory)

    try:
        ts_df1 = pd.read_csv(troublefile, low_memory=False)
        ts_df1 = ts_df1[['Código', 'Causas', 'Recomendações', 'Peso']]
        ts_df1.columns = ['Code', 'Causas', 'Recomendações', 'Peso']
        ts_df1.dropna(how='all', inplace=True)
    except FileNotFoundError:
        exit(0)

    checkdestiny(workfolder)
    try:
        cdf = pd.read_csv(destinationfolder + '/events_output.csv', low_memory=False)
        cdf.to_csv(workfolder + '/events_output.csv', encoding='utf-8-sig', index=False)
    except FileNotFoundError:
        cdf = pd.DataFrame(columns=std_event_list)
        cdf.to_csv(workfolder + '/events_output.csv', encoding='utf-8-sig', index=False)

    checkdestiny(workfolder2)
    try:
        cdf1 = pd.read_csv(workfolder2 + '/eventssum_output.csv', low_memory=False)
    except FileNotFoundError:
        cdf1 = pd.DataFrame(columns=std_eventsum_list)
        cdf1.to_csv(workfolder2 + '/eventssum_output.csv', encoding='utf-8-sig', index=False)

    if eventfile.endswith('.xlsx'):
        wb = openpyxl.load_workbook(eventfile)
        print(' ')
        print('Iniciando tratamento de dados de Eventos e Alertas...')
        print(' ')
        for a_name in asset_list:
            if a_name != 'Totals':
                asset_sn = a_name[-8:]
                outws = workfolder + '/' + asset_sn + '.csv'
                outws2 = workfolder2 + '/' + asset_sn + '.csv'
                try:
                    ws = wb[str(a_name)]
                    print(a_name, 'Ok')
                    rotinaseventos(outws, ws, asset_sn)
                    rotinaseventossum(outws2, ws, asset_sn)
                except KeyError:
                    print(a_name, 'Vazio')
                    ws = pd.DataFrame(columns=std_event_list)
                    ws['Asset'] = asset_sn
                    ws['Site'] = findsitename(asset_sn)
                    ws['Timestamp'] = pd.to_datetime(ws['Timestamp'])
                    ws['Code'] = ws['Code'].astype(str)
                    ws.to_csv(outws, encoding='utf-8-sig', index=False)

                    ws2 = pd.DataFrame(columns=std_eventsum_list)
                    ws2['Asset'] = asset_sn
                    ws2['Site'] = findsitename(asset_sn)
                    ws2['Code'] = ws2['Code'].astype(str)
                    ws2.to_csv(outws2, encoding='utf-8-sig', index=False)

        # CONCATENA TUDO NA PLANILHA FINAL
    concatenarev(workfolder, 'events_')
    evdfoutput = concatenar_profile(workfolder2, std_eventsum_list, 'eventssum_')

    print(' ')
    print('Iniciando aquisição de texto de troubleshoot e cálculo do índice de confiabilidade...')
    print(' ')
    # MANUTENÇÃO COMEÇA CALCULANDO CONSUMO POR DIA, TOTAL DE CONSUMO, HORAS DE SERVIÇO POR DIA, TOTAL DE HORAS DE SERVIÇO

    ######  ÍNDICE DE CONFIABILIDADE

    evdfoutput['Code'] = evdfoutput['Code'].str.replace(':', '--')
    evdfoutput.loc[evdfoutput['Type'] == 'Event', 'IsEv'] = 'E'
    evdfoutput['CodeB'] = evdfoutput['Code'].replace("--.+", "", regex=True)
    evdfoutput['CodeB'] = evdfoutput['IsEv'] + evdfoutput['CodeB']
    evdfoutput.loc[evdfoutput['Type'] == 'Event', 'Code'] = evdfoutput.loc[evdfoutput['Type'] == 'Event', 'CodeB']

    if keepfiles == 0:
        evdfoutput.drop(columns=['IsEv', 'CodeB'], inplace=True)

    # ESSE É O CERTO: ts_df1.replace("CONTRABARRA(.+", "", inplace=True, regex=True)
    ts_df1.replace("\(.+", "", inplace=True, regex=True)

    evdfoutput.dropna(how='all', inplace=True)
    evdfoutput = pd.merge_ordered(evdfoutput, ts_df1, how='left')
    evdfoutput['PesoSev'] = evdfoutput['Severity']

    evdfoutput['PesoSev'] = evdfoutput['PesoSev'].str.replace('Low', '5.6234')
    evdfoutput['PesoSev'] = evdfoutput['PesoSev'].str.replace('Medium', '10')
    evdfoutput['PesoSev'] = evdfoutput['PesoSev'].str.replace('High', '17.7828')

    evdfoutput['PesoSev'] = pd.to_numeric(evdfoutput['PesoSev'])
    evdfoutput['Peso'] = pd.to_numeric(evdfoutput['Peso'])

    evdfoutput['PesoReal'] = np.power(evdfoutput['PesoSev'], 4) * np.power(evdfoutput['Peso'], 2)
    evdfoutput['IC'] = ((evdfoutput['Count'] * evdfoutput['PesoReal']) / 10) / Pdias

    # ---
    # Definição do status de confiabilidade
    sconf2 = 133
    sconf3 = 333
    evdfoutput["Status de confiabilidade"] = "Verde"

    evdfoutput.loc[evdfoutput['IC'] >= sconf2, 'Status de confiabilidade'] = 'Amarelo'
    evdfoutput.loc[evdfoutput['IC'] >= sconf3, 'Status de confiabilidade'] = 'Vermelho'

    # ---
    outfs = workfolder2 + '/eventssum_output.csv'
    if keepfiles == 0:
        evdfoutput.drop(columns=['PesoSev', 'Peso', 'PesoReal'], inplace=True)

    evdfoutput.to_csv(outfs, encoding='utf-8-sig', index=False)
    txt = open(destinationfolder + "/Events_html.txt", "w+")
    txt.write(evhtml(evdfoutput))
    txt.close()

def evhtml(df: pd.core.frame.DataFrame):
    """ 
        Função que gera código em formatação HTML dos alertas.
  
        Parameters: 
            df (DataFrame): Caminho do arquivo Excel com os alertas que a serem descritos
            
        Returns: 
            Txhtml (srt): Descrição de alertas em HTML
        """
    df = df[["Code", "Description", "Causas", "Recomendações", ]].drop_duplicates(subset=["Code"])
    df.reset_index(inplace=True)
    Txhtml = ""
    i = -1

    for VNome in df["Description"]:
        i += 1
        t1 = VNome
        t2 = df.loc[i, 'Causas']
        t3 = df.loc[i, 'Recomendações']
        Txhtml = Txhtml + "<b>" + str(t1) + "</b><br>" + "<b>Causas possíveis: </b>" + str(
            t2) + "<br><b>Recomendações: </b>" + str(t3) + "<br><br>"
    return Txhtml.replace('\n', '')

#####################################################################################################
#                               CONVERSÃO DE HISTÓRICOS DE DADOS                                    #
#####################################################################################################

def historyconvert(historyfile):
    global asset_sn
    global dfoutput
    global fcd, fct, shd, sht
    global dfo, dfm
    dfm = pd.DataFrame()

    historyfile = os.path.join(historyfile)

    workfolder = os.path.join(destinationfolder, historydirectory)

    checkdestiny(workfolder)
    try:
        cdf = pd.read_csv(destinationfolder + '/history_output.csv', low_memory=False)
        cdf.to_csv(workfolder + '/history_output.csv', encoding='utf-8-sig', index=False)
        
    except FileNotFoundError:
        cdf = pd.DataFrame(columns=std_param_list)
        cdf.to_csv(workfolder + '/history_output.csv', encoding='utf-8-sig', index=False)

    dailyfolder = os.path.join(destinationfolder, dailysumdir)
    checkdestiny(dailyfolder)

    try:
        cdf = pd.read_csv(destinationfolder + '/historyday_output.csv', low_memory=False)
        cdf.to_csv(dailyfolder + '/historyday_output.csv', encoding='utf-8-sig', index=False)
    except FileNotFoundError:
        cdf = pd.DataFrame(columns=std_param_list)
        cdf.to_csv(dailyfolder + '/historyday_output.csv', encoding='utf-8-sig', index=False)
    #############################################################################
    ##################   CASO O ARQUIVO DE ENTRADA SEJA XLSX   ##################
    #############################################################################

    if historyfile.endswith('.xlsx'):
        global ws
        global Pdias
        wb = openpyxl.load_workbook(historyfile)
        print(' ')
        print('Iniciando Extracting, Transforming and Loading...')
        print(' ')

        for a_name in asset_list:
            if a_name != 'Totals':
                asset_sn = a_name[-8:]
                outws = workfolder + '/' + asset_sn + '.csv'
                try:
                    ws = wb[str(a_name)]
                    wsdata = ws.values
                    cols = next(wsdata)[0:]
                    dataframe = pd.DataFrame(wsdata, columns=cols)
                    dataframe = replace_text(dataframe, data_replace_text)
                    dataframe['Timestamp'] = pd.to_datetime(dataframe['Timestamp'])
                except KeyError:
                    dataframe = pd.DataFrame(columns=std_param_list)
                    dataframe['Timestamp'] = pd.to_datetime(dataframe['Timestamp'])
                    dataframe.to_csv(outws, encoding='utf-8-sig', index=False)

                rotinas(dataframe)
                # salva dataframe pronto na pasta dos parametros
                dfclean.to_csv(os.path.join(destinationfolder, workfolder) + asset_sn + '.csv', encoding='utf-8-sig',
                               index=False)

                dfhday = dfclean
                dfhday.reset_index(drop=True, inplace=True)
                dfhday.set_index(pd.DatetimeIndex(dfhday['Timestamp']), inplace=True)
                dfhday = dfhday.resample('D').mean(numeric_only=True)
                dfhday = dfhday.replace(0, np.nan)
                dfhday['Asset'] = asset_sn
                dfhday['Site'] = findsitename(asset_sn)
                dailyfile = os.path.join(dailyfolder) + '/' + asset_sn + '.csv'
                dfhday.to_csv(dailyfile, encoding='utf-8-sig', index=True)

        # CONCATENA TUDO NA PLANILHA FINAL
        dfoutput = concatenar(os.path.join(destinationfolder, workfolder), 'history_')

        Pdias = np.timedelta64(
            pd.to_datetime(dfoutput['Timestamp']).max() - pd.to_datetime(dfoutput['Timestamp']).min(), 'D').astype(int)

        concatenar_profile(os.path.join(destinationfolder, loadhistdir), std_loadhist_list, 'loadhist_')
        concatenar_profile(os.path.join(destinationfolder, rpmhistdir), std_rpmhist_list, 'rpmhist_')
        #concatenar_study(dailyfolder, 'historyday_')
        concatenar(dailyfolder, 'historyday_')

        Pdias = np.timedelta64(
            pd.to_datetime(dfoutput['Timestamp']).max() - pd.to_datetime(dfoutput['Timestamp']).min(), 'D').astype(int)

        print(' ')
        print('Iniciando calculos de intervalos de manutenção...')
        print(' ')
        # MANUTENÇÃO COMEÇA CALCULANDO CONSUMO POR DIA, TOTAL DE CONSUMO, HORAS DE SERVIÇO POR DIA, TOTAL DE HORAS DE SERVIÇO
        dfo = pd.DataFrame(columns=maintenance_list)

        for a_name in asset_list:
            if a_name != 'Totals':
                asset_sn = a_name[-8:]
                dfa = dfoutput.query('Asset == @asset_sn')
                lastused = pd.to_datetime(dfa['Timestamp']).max()
                datasetvazio = 0
                if dfa.empty:
                    datasetvazio = 1
                # print(asset_sn + ' - Dataset Vazio: ' + str(datasetvazio))
                dfm = maintenanceoutput(dfoutput, lastused, asset_sn, datasetvazio)
                dfo = pd.concat([dfo, dfm])
                dfo.drop_duplicates(subset=None, inplace=True, keep='last')
                dfo.dropna(how='all', axis=1, inplace=True)
        checkdestiny(os.path.join(destinationfolder, mandirectory))
        dfo.to_csv(os.path.join(destinationfolder, mandirectory) + 'maintenance_output.csv', encoding='utf-8-sig',
                   index=False)

        ####### UTILIZAÇÃO #######
        print(' ')
        print('Iniciando estudos de Utilização e Taxas de Consumo de Combustível...')
        print(' ')

        outdir1 = os.path.join(destinationfolder, loadstudydir)
        outdir1h = os.path.join(destinationfolder, loadstudydirh)
        outdir2 = os.path.join(destinationfolder, loadresumedir)
        outdir3 = os.path.join(destinationfolder, fuelstudydir)
        outdir4 = os.path.join(destinationfolder, fuelresumedir)
        checkdestiny(outdir1)
        checkdestiny(outdir1h)
        checkdestiny(outdir2)
        checkdestiny(outdir3)
        checkdestiny(outdir4)

        for sit in sites_list:
            print('Sítio:', sit)
            slist = findsiteassets(sit)
            count = 0
            salist = []
            for a_name in slist:
                if count < 1:
                    df1 = dfoutput.query('Asset == @a_name')[
                        ['Timestamp', 'Load', 'Fuel_Rate', 'Total_Fuel_DIFF', 'SMH_DIFF', 'EXH_DIFF']]
                    df1.columns = ['Timestamp' if x == 'Timestamp' else str(a_name) + '_' + x for x in df1.columns]
                    df1.reset_index(drop=True, inplace=True)
                    dfh = df1
                    salist.append(a_name)
                else:
                    df1 = dfoutput.query('Asset == @a_name')[
                        ['Timestamp', 'Load', 'Fuel_Rate', 'Total_Fuel_DIFF', 'SMH_DIFF', 'EXH_DIFF']]
                    df1.columns = ['Timestamp' if x == 'Timestamp' else str(a_name) + '_' + x for x in df1.columns]
                    df1.reset_index(drop=True, inplace=True)
                    if 'Timestamp' in dfh.columns:
                        dfh = pd.merge_ordered(dfh, df1, fill_method='None')
                    salist.append(a_name)
                count += 1

            # dfh.dropna(how='all', axis=1, inplace=True)
            dfh.reset_index(drop=True, inplace=True)
            dfh.set_index(pd.DatetimeIndex(dfh['Timestamp']), inplace=True)

            dfhhr = dfh.resample('1H').mean(numeric_only=True)
            dfhhr = dfhhr.replace(0, np.nan)

            dfhmin = dfh.resample('5Min').mean(numeric_only=True)
            dfhmin = dfhmin.replace(0, np.nan)

            print('Lista de Ativos:', salist)

            # LOAD RESUME
            loadtable = pd.concat([dfhhr.loc[:, dfhhr.columns.isin(['Timestamp'])],
                                   dfhhr.loc[:, dfhhr.columns.str.contains('Load')]], axis=1)
            loadtable['QNT_SIM'] = loadtable.apply(lambda x: x.notnull().sum(), axis='columns')
            
            fin = pd.concat([pd.DataFrame([i], columns=['QNT']) for i in range(len(slist) + 1)], ignore_index=True)
            l = []
            for i in range(len(slist) + 1):
                l.append(loadtable[loadtable['QNT_SIM'] == (i)].count()['QNT_SIM'] / len(loadtable['QNT_SIM']) * 100)
            fin['%'] = l
            fin['Site'] = sit
            finfile = os.path.join(outdir2) + '/' + sit + '_LOAD_RESUME' + '.csv'
            fin.to_csv(finfile, encoding='utf-8-sig', index=False)

            # LOAD STUDY
            ltvd = loadtable.resample('1D').apply(np.ceil).round(0)
            for x in range(len(loadtable.columns) - 1):
                ltvd['UT' + str(x + 1) + 'E'] = (
                        (loadtable['QNT_SIM'] == (x + 1)).resample('1D').sum().astype(int) / 24 * 100).round(1)
            if keepfiles == 0:
                ltvd.drop(['QNT_SIM'], axis=1, inplace=True)
            ltvd['Site'] = sit
            finfile2 = os.path.join(outdir1) + '/' + sit + '_LOAD_STUDY' + '.csv'
            ltvd.to_csv(finfile2, encoding='utf-8-sig', index=True)

            # LOAD STUDY 1h

            loadtablemin = pd.concat([dfhmin.loc[:, dfhmin.columns.isin(['Timestamp'])],
                                      dfhmin.loc[:, dfhmin.columns.str.contains('Load')]], axis=1)

            loadtablemin['QNT_SIM'] = loadtablemin.apply(lambda x: x.notnull().sum(), axis='columns')

            ltvdh = loadtablemin.resample('1H').mean(numeric_only=True).round(0)
            for x in range(len(loadtablemin.columns) - 1):
                ltvdh['UT' + str(x + 1) + 'E'] = (
                        (loadtablemin['QNT_SIM'] == (x + 1)).resample('1H').sum().astype(int) / 12 * 100).round(1)
            if keepfiles == 0:
                ltvdh.drop(['QNT_SIM'], axis=1, inplace=True)
            ltvdh['Site'] = sit
            finfile2h = os.path.join(outdir1h) + '/' + sit + '_LOAD_STUDY_H' + '.csv'
            ltvdh.to_csv(finfile2h, encoding='utf-8-sig', index=True)

            # FUEL RESUME
            fratetable = pd.concat([dfhhr.loc[:, dfhhr.columns.isin(['Timestamp'])],
                                    dfhhr.loc[:, dfhhr.columns.str.contains('Fuel_Rate')]], axis=1)

            rtvb = fratetable.copy()
            rtvb['AVG'] = rtvb.sum(numeric_only=True, axis=1)

            rtvfile = os.path.join(outdir3) + '/' + sit + '_FR_STUDY' + '.csv'
            fratetable = fratetable.replace(0, np.nan)
            rtvD = fratetable.resample('1D').mean(numeric_only=True).round(0)
            rtvD['Site'] = sit
            rtvD.to_csv(rtvfile, encoding='utf-8-sig', index=True)

            fratetable['QNT_SIM'] = fratetable.apply(lambda x: x.notnull().sum(), axis='columns')

            rin = pd.concat([pd.DataFrame([i], columns=['QNT']) for i in range(len(slist) + 1)], ignore_index=True)
            l = []
            for i in range(len(slist) + 1):
                l.append(rtvb[fratetable['QNT_SIM'] == (i)].mean(numeric_only=True)['AVG'])
            rin['L/hr'] = l
            rin['Site'] = sit
            rinfile = os.path.join(outdir4) + '/' + sit + '_FR_RESUME' + '.csv'
            rin.to_csv(rinfile, encoding='utf-8-sig', index=False)

        dfs_load = concatenar_study(outdir1, 'loadstudy_')
        dfs_loadh = concatenar_study(outdir1h, 'loadstudy_H_')
        concatenar_study(outdir2, 'loadresume_')
        dfs_fr = concatenar_study(outdir3, 'fuelstudy_')
        concatenar_study(outdir4, 'fuelresume_')
       
        dfs_load.to_csv(destinationfolder + 'testea.csv',encoding='utf-8-sig', index=False)
        dfs_loadh.to_csv(destinationfolder + 'testeb.csv',encoding='utf-8-sig', index=False)
        dfs_fr.to_csv(destinationfolder + 'testec.csv',encoding='utf-8-sig', index=False)

    ############################################################################
    ##################   CASO O ARQUIVO DE ENTRADA SEJA ZIP   ##################
    ############################################################################
    elif historyfile.endswith('.zip'):
        zf = zipfile.ZipFile(historyfile)
        print(' ')
        print('Iniciando Extracting, Transforming and Loading...')
        print(' ')

        for a_name in asset_list:
            if a_name != 'Totals':
                ws = a_name + '.csv'

                asset_sn = a_name[-8:]
                try:
                    dataframe = pd.read_csv(zf.open(ws), encoding='utf-16le', dtype=object)
                    dataframe = replace_text(dataframe, data_replace_text)
                    dataframe['Timestamp'] = pd.to_datetime(dataframe['Timestamp'])

                except KeyError:
                    dataframe = pd.DataFrame(columns=std_param_list)
                    dataframe['Timestamp'] = pd.to_datetime(dataframe['Timestamp'])

                # chama rotinas aplicáveis ao dataframe
                rotinas(dataframe)
                # salva dataframe pronto na pasta dos parametros
                dfclean.to_csv(os.path.join(destinationfolder, workfolder) + asset_sn + '.csv', encoding='utf-8-sig',
                               index=False)

                dfhday = dfclean
                dfhday.reset_index(drop=True, inplace=True)
                dfhday.set_index(pd.DatetimeIndex(dfhday['Timestamp']), inplace=True)
                dfhday = dfhday.resample('D').mean(numeric_only=True)
                dfhday = dfhday.replace(0, np.nan)
                dfhday['Asset'] = asset_sn
                dfhday['Site'] = findsitename(asset_sn)
                dailyfile = os.path.join(dailyfolder) + '/' + asset_sn + '.csv'
                dfhday.to_csv(dailyfile, encoding='utf-8-sig', index=True)

        # CONCATENA TUDO NA PLANILHA FINAL
        dfoutput = concatenar(os.path.join(destinationfolder, workfolder), 'history_')

        Pdias = np.timedelta64(
            pd.to_datetime(dfoutput['Timestamp']).max() - pd.to_datetime(dfoutput['Timestamp']).min(), 'D').astype(int)

        concatenar_profile(os.path.join(destinationfolder, loadhistdir), std_loadhist_list, 'loadhist_')
        concatenar_profile(os.path.join(destinationfolder, rpmhistdir), std_rpmhist_list, 'rpmhist_')
        concatenar(dailyfolder, 'historyday_')

        print(' ')
        print('Iniciando calculos de intervalos de manutenção...')
        print(' ')
        # MANUTENÇÃO COMEÇA CALCULANDO CONSUMO POR DIA, TOTAL DE CONSUMO, HORAS DE SERVIÇO POR DIA, TOTAL DE HORAS DE SERVIÇO
        dfo = pd.DataFrame(columns=maintenance_list)
        for a_name in asset_list:
            if a_name != 'Totals':
                asset_sn = a_name[-8:]
                dfa = dfoutput.query('Asset == @asset_sn')
                lastused = pd.to_datetime(dfa['Timestamp']).max()
                datasetvazio = 0
                if dfa.empty:
                    datasetvazio = 1
                # print(asset_sn + ' - Dataset Vazio: ' + str(datasetvazio))
                dfm = maintenanceoutput(dfoutput, lastused, asset_sn, datasetvazio)
                dfo = pd.concat([dfo, dfm])
                dfo.drop_duplicates(subset=None, inplace=True, keep='last')
                dfo.dropna(how='all', axis=1, inplace=True)
        checkdestiny(os.path.join(destinationfolder, mandirectory))
        dfo.to_csv(os.path.join(destinationfolder, mandirectory) + 'maintenance_output.csv', encoding='utf-8-sig',
                   index=False)

        ####### UTILIZAÇÃO #######
        print(' ')
        print('Iniciando estudos de Utilização e Taxas de Consumo de Combustível...')
        print(' ')

        outdir1 = os.path.join(destinationfolder, loadstudydir)
        outdir1h = os.path.join(destinationfolder, loadstudydirh)
        outdir2 = os.path.join(destinationfolder, loadresumedir)
        outdir3 = os.path.join(destinationfolder, fuelstudydir)
        outdir4 = os.path.join(destinationfolder, fuelresumedir)
        checkdestiny(outdir1)
        checkdestiny(outdir1h)
        checkdestiny(outdir2)
        checkdestiny(outdir3)
        checkdestiny(outdir4)

        for sit in sites_list:
            print('Sítio:', sit)
            slist = findsiteassets(sit)
            count = 0
            salist = []
            for a_name in slist:
                if count < 1:
                    df1 = dfoutput.query('Asset == @a_name')[['Timestamp', 'Load', 'Fuel_Rate']]
                    df1.columns = ['Timestamp' if x == 'Timestamp' else str(a_name) + '_' + x for x in df1.columns]
                    df1.reset_index(drop=True, inplace=True)
                    dfh = df1
                    salist.append(a_name)
                else:
                    df1 = dfoutput.query('Asset == @a_name')[['Timestamp', 'Load', 'Fuel_Rate']]
                    df1.columns = ['Timestamp' if x == 'Timestamp' else str(a_name) + '_' + x for x in df1.columns]
                    df1.reset_index(drop=True, inplace=True)
                    if 'Timestamp' in dfh.columns:
                        dfh = pd.merge_ordered(dfh, df1, fill_method='None')
                    salist.append(a_name)
                count += 1

            # dfh.dropna(how='all', axis=1, inplace=True)
            dfh.reset_index(drop=True, inplace=True)
            dfh.set_index(pd.DatetimeIndex(dfh['Timestamp']), inplace=True)

            dfhhr = dfh.resample('1H').mean(numeric_only=True)
            dfhhr = dfhhr.replace(0, np.nan)

            dfhmin = dfh.resample('5Min').mean(numeric_only=True)
            dfhmin = dfhmin.replace(0, np.nan)

            print('Lista de Ativos:', salist)

            # LOAD RESUME
            loadtable = pd.concat([dfhhr.loc[:, dfhhr.columns.isin(['Timestamp'])],
                                   dfhhr.loc[:, dfhhr.columns.str.contains('Load')]], axis=1)

            loadtable['QNT_SIM'] = loadtable.apply(lambda x: x.notnull().sum(), axis='columns')

            fin = pd.concat([pd.DataFrame([i], columns=['QNT']) for i in range(len(slist) + 1)], ignore_index=True)
            l = []
            for i in range(len(slist) + 1):
                l.append(loadtable[loadtable['QNT_SIM'] == (i)].count()['QNT_SIM'] / len(loadtable['QNT_SIM']) * 100)
            fin['%'] = l
            fin['Site'] = sit
            finfile = os.path.join(outdir2) + '/' + sit + '_LOAD_RESUME' + '.csv'
            fin.to_csv(finfile, encoding='utf-8-sig', index=False)

            # LOAD STUDY
            ltvd = loadtable.resample('1D').mean(numeric_only=True).round(0)
            for x in range(len(loadtable.columns) - 1):
                ltvd['UT' + str(x + 1) + 'E'] = (
                            (loadtable['QNT_SIM'] == (x + 1)).resample('1D').sum().astype(int) / 24 * 100).round(1)
            if keepfiles == 0:
                ltvd.drop(['QNT_SIM'], axis=1, inplace=True)
            ltvd['Site'] = sit
            finfile2 = os.path.join(outdir1) + '/' + sit + '_LOAD_STUDY' + '.csv'
            ltvd.to_csv(finfile2, encoding='utf-8-sig', index=True)

            # LOAD STUDY 1h
            loadtablemin = pd.concat([dfhmin.loc[:, dfhmin.columns.isin(['Timestamp'])],
                                      dfhmin.loc[:, dfhmin.columns.str.contains('Load')]], axis=1)
            loadtablemin['QNT_SIM'] = loadtablemin.apply(lambda x: x.notnull().sum(), axis='columns')

            ltvdh = loadtablemin.resample('1H').mean(numeric_only=True).round(0)
            for x in range(len(loadtablemin.columns) - 1):
                ltvdh['UT' + str(x + 1) + 'E'] = (
                            (loadtablemin['QNT_SIM'] == (x + 1)).resample('1H').sum().astype(int) / 12 * 100).round(1)
            if keepfiles == 0:
                ltvdh.drop(['QNT_SIM'], axis=1, inplace=True)
            ltvdh['Site'] = sit
            finfile2h = os.path.join(outdir1h) + '/' + sit + '_LOAD_STUDY_H' + '.csv'
            ltvdh.to_csv(finfile2h, encoding='utf-8-sig', index=True)

            # FUEL RESUME
            fratetable = pd.concat([dfhhr.loc[:, dfhhr.columns.isin(['Timestamp'])],
                                    dfhhr.loc[:, dfhhr.columns.str.contains('Fuel_Rate')]], axis=1)

            rtvb = fratetable.copy()
            rtvb['AVG'] = rtvb.sum(numeric_only=True, axis=1)

            rtvfile = os.path.join(outdir3) + '/' + sit + '_FR_STUDY' + '.csv'
            fratetable = fratetable.replace(0, np.nan)
            rtvD = fratetable.resample('1D').mean(numeric_only=True).round(0)
            rtvD['Site'] = sit
            rtvD.to_csv(rtvfile, encoding='utf-8-sig', index=True)

            fratetable['QNT_SIM'] = fratetable.apply(lambda x: x.notnull().sum(), axis='columns')

            rin = pd.concat([pd.DataFrame([i], columns=['QNT']) for i in range(len(slist) + 1)], ignore_index=True)
            l = []
            for i in range(len(slist) + 1):
                l.append(rtvb[fratetable['QNT_SIM'] == (i)].mean(numeric_only=True)['AVG'])
            rin['L/hr'] = l
            rin['Site'] = sit
            rinfile = os.path.join(outdir4) + '/' + sit + '_FR_RESUME' + '.csv'
            rin.to_csv(rinfile, encoding='utf-8-sig', index=False)

        concatenar_study(outdir1, 'loadstudy_')
        concatenar_study(outdir1h, 'loadstudy_H_')
        concatenar_study(outdir2, 'loadresume_')
        concatenar_study(outdir3, 'fuelstudy_')
        concatenar_study(outdir4, 'fuelresume_')

#####################################################################################################
#                                           ENVIROMENT                                              #
#####################################################################################################

def preplistas(engine_file, event_file, ts_file, destfolder, deb):

    global scriptname
    scriptname = os.path.basename(__file__)

    global swdir
    # swdir = os.getcwd()
    swdir = (Path(destfolder).parent)
    print(swdir)

    global destinationfolder
    destinationfolder = destfolder

    global keepfiles
    keepfiles = deb

    global historydirectory
    historydirectory = 'History/'
    global eventsdirectory
    eventsdirectory = 'Events/'
    global eventssumdirectory
    eventssumdirectory = 'EventsSummary/'
    global mandirectory
    mandirectory = 'Maintenance/'
    global rpmhistdir
    rpmhistdir = 'RPMProfile/'
    global loadhistdir
    loadhistdir = 'LoadProfile/'
    global loadresumedir
    loadresumedir = 'LoadResume/'
    global fuelresumedir
    fuelresumedir = 'FuelResume/'
    global loadstudydir
    loadstudydir = 'LoadStudy/'
    global loadstudydirh
    loadstudydirh = 'LoadStudy_H/'
    global fuelstudydir
    fuelstudydir = 'FuelStudy/'
    global dailysumdir
    dailysumdir = 'DailySummary/'

    global infodir
    infodir = os.path.join(swdir, '00_INFOS/')
    assetinfofile = infodir + 'ASSET_INFO.xlsx'
    global assetlistdf
    assetlistdf = openfilewb(assetinfofile,'ASSET_LIST')

    global asset_list
    asset_list = getlistativos(event_file)
    global sites_list
    sites_list = getlistasites(assetinfofile)

    global f_density
    f_density = 850

    destinationfolder = os.path.join(destinationfolder)
    checkdestiny(destinationfolder)
    historyconvert(engine_file)
    eventsconvert(event_file, ts_file)
    limpadao(destinationfolder)

    print('Finished RFV 2.0 Handling Script v4')
    print()


# if __name__ == '__main__':
#     print('Você deve executar o arquivo GUI.py do conversor.')
#     exit(0)

########################################################################################
#                                                                                      #
#                                     TESTE                                            #
#                                                                                      #
########################################################################################

# engine_file="D:/DEV/EquipmentAnalytics/02_DATALOGS/00_Jun2021_Apr2022/HistoricalData-Transposed-2022-04-29.zip"
# event_file="D:/DEV/EquipmentAnalytics/02_DATALOGS/00_Jun2021_Apr2022/EngineEventHistory-2022-04-29.xlsx"
# ts_file="D:/DEV/EquipmentAnalytics/TROUBLESHOOT.csv"
# path="D:/DEV/EquipmentAnalytics/01_BD_CUSTOMER"
# deb=1

# preplistas(engine_file,event_file,ts_file,path,deb)